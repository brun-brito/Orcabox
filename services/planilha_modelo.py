import openpyxl
from openpyxl.utils import get_column_letter

def criar_planilha_modelo(file_path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Produtos"

    # Definindo os cabeçalhos das colunas
    colunas = ['Nome', 'Preço', 'Quantidade', 'Marca', 'Categoria']
    for idx, coluna in enumerate(colunas, 1):
        letra_coluna = get_column_letter(idx)
        sheet[f'{letra_coluna}1'] = coluna

    # Congela a primeira linha para que o usuário não possa editá-la
    sheet.freeze_panes = "A2"

    # Define a largura das colunas
    for idx, _ in enumerate(colunas, 1):
        sheet.column_dimensions[get_column_letter(idx)].width = 20

    # Salva a planilha modelo
    workbook.save(file_path)
